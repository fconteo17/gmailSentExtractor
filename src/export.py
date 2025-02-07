"""
Export operations for the Gmail Export Tool.
"""
import logging
from typing import List, Dict
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class ExportManager:
    """Handles export operations for email data."""

    def __init__(self):
        """Initialize the export manager."""
        self.logger = logging.getLogger(__name__)

    def _clean_date(self, date_str: str) -> str:
        """Clean and validate date string."""
        if not date_str or date_str == "No Date":
            return "1900-01-01 00:00:00"  # Default date for invalid entries
        try:
            # Try to parse and reformat the date
            dt = pd.to_datetime(date_str)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except:
            return "1900-01-01 00:00:00"  # Default date for parsing errors

    def _validate_data(self, emails_data: list) -> list:
        """Validate and clean email data before export."""
        cleaned_data = []
        for email in emails_data:
            cleaned_email = email.copy()
            # Clean date
            cleaned_email['Date'] = self._clean_date(email.get('Date', 'No Date'))
            
            # Construct full email from Username and Domain
            username = email.get('Username', '')
            domain = email.get('Domain', '')
            if username and domain:
                cleaned_email['Email'] = f"{username}@{domain}"
            else:
                cleaned_email['Email'] = "No Email"
            
            # Keep the original fields
            cleaned_email['Domain'] = domain if domain else "No Domain"
            cleaned_email['Subject'] = email.get('Subject', 'No Subject')
            
            cleaned_data.append(cleaned_email)
        return cleaned_data

    def export_to_excel(self, emails_data: list, output_file: str, email: str) -> bool:
        """
        Export email data to Excel file.
        
        Args:
            emails_data: List of email data dictionaries.
            output_file: Path to output Excel file.
            email: Email address being exported.
            
        Returns:
            bool: True if successful, False otherwise.
        """
        try:
            if not emails_data:
                self.logger.error("No data to export")
                return False

            # Ensure the export directory exists
            os.makedirs(os.path.dirname(output_file), exist_ok=True)

            # Clean and validate data
            cleaned_data = self._validate_data(emails_data)
            
            # Convert to DataFrame
            df = pd.DataFrame(cleaned_data)
            
            # Ensure all dates are in datetime format
            df["Date"] = pd.to_datetime(df["Date"], format="%Y-%m-%d %H:%M:%S", errors='coerce')
            
            # Remove any rows where date conversion failed
            df = df.dropna(subset=["Date"])
            
            # Sort by date
            df = df.sort_values("Date")
            
            # Convert back to string format for Excel
            df["Date"] = df["Date"].dt.strftime("%Y-%m-%d %H:%M:%S")

            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Sent Emails - {email}"

            # Write headers
            columns = ["Date", "Email", "Domain", "Subject"]
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Write data
            for row_idx, row in enumerate(df[columns].values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            # Save the workbook
            if os.path.exists(output_file):
                os.remove(output_file)  # Remove existing file to avoid corruption
            wb.save(output_file)
            wb.close()  # Explicitly close the workbook
            
            self.logger.info(f"Successfully exported {len(emails_data)} emails to {output_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {str(e)}")
            return False 